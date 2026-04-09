#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File Converter Service
处理各种文件格式转换（HEIC/OFD/PDF等）
"""

import os
import io
import base64
import tempfile
from flask import Flask, request, jsonify, send_file
from PIL import Image

app = Flask(__name__)

# 检测支持的格式转换
SUPPORTED_FORMATS = {
    'image': ['jpg', 'jpeg', 'png', 'bmp', 'webp', 'gif', 'tiff'],
    'heic': ['heic', 'heif'],
    'pdf': ['pdf'],
    'ofd': ['ofd'],
    'document': ['docx', 'xlsx']
}

# 检查依赖
HAS_HEIF = False
HAS_PDF2IMAGE = False
HAS_PYPDF = False
HAS_DOCX = False
HAS_OPENPYXL = False

try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
    HAS_HEIF = True
except ImportError:
    pass

try:
    from pdf2image import convert_from_path
    HAS_PDF2IMAGE = True
except ImportError:
    pass

try:
    from pypdf import PdfReader
    HAS_PYPDF = True
except ImportError:
    pass

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    pass

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    pass


@app.route('/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'formats': {
            'heic': HAS_HEIF,
            'pdf': HAS_PDF2IMAGE or HAS_PYPDF,
            'docx': HAS_DOCX,
            'xlsx': HAS_OPENPYXL
        }
    })


@app.route('/convert', methods=['POST'])
def convert_file():
    """
    文件转换接口
    请求格式:
    {
        "file": "base64编码的文件数据",
        "filename": "原始文件名",
        "target_format": "目标格式（jpg/png等）"
    }
    """
    try:
        data = request.get_json()
        if not data or 'file' not in data:
            return jsonify({'error': '缺少文件数据'}), 400

        file_data = data['file']
        if file_data.startswith('data:'):
            file_data = file_data.split(',', 1)[1]

        filename = data.get('filename', 'unknown')
        target_format = data.get('target_format', 'jpg').lower()

        # 解码文件
        file_bytes = base64.b64decode(file_data)

        # 根据文件扩展名选择转换方法
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

        if ext in SUPPORTED_FORMATS['heic']:
            return convert_heic(file_bytes, target_format)
        elif ext == 'pdf':
            return convert_pdf(file_bytes, target_format)
        elif ext == 'ofd':
            return convert_ofd(file_bytes, target_format)
        elif ext == 'docx':
            return convert_docx(file_bytes, target_format)
        elif ext == 'xlsx':
            return convert_xlsx(file_bytes, target_format)
        elif ext in SUPPORTED_FORMATS['image']:
            return convert_image(file_bytes, target_format)
        else:
            return jsonify({'error': f'不支持的文件格式: {ext}'}), 400

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def convert_image(file_bytes, target_format='jpg'):
    """转换普通图片格式"""
    try:
        image = Image.open(io.BytesIO(file_bytes))

        # 转换颜色模式
        if target_format.lower() in ['jpg', 'jpeg']:
            if image.mode in ('RGBA', 'P'):
                image = image.convert('RGB')

        # 输出
        output = io.BytesIO()
        image.save(output, format=target_format.upper())
        output.seek(0)

        return jsonify({
            'success': True,
            'image': base64.b64encode(output.read()).decode('utf-8'),
            'format': target_format
        })

    except Exception as e:
        return jsonify({'error': f'图片转换失败: {str(e)}'}), 500


def convert_heic(file_bytes, target_format='jpg'):
    """转换HEIC/HEIF格式"""
    if not HAS_HEIF:
        return jsonify({'error': 'HEIC格式支持未安装，请安装: pip install pillow-heif'}), 500

    try:
        image = Image.open(io.BytesIO(file_bytes))

        if target_format.lower() in ['jpg', 'jpeg']:
            if image.mode in ('RGBA', 'P'):
                image = image.convert('RGB')

        output = io.BytesIO()
        image.save(output, format=target_format.upper())
        output.seek(0)

        return jsonify({
            'success': True,
            'image': base64.b64encode(output.read()).decode('utf-8'),
            'format': target_format
        })

    except Exception as e:
        return jsonify({'error': f'HEIC转换失败: {str(e)}'}), 500


def convert_pdf(file_bytes, target_format='jpg'):
    """转换PDF为图片"""
    if not HAS_PDF2IMAGE:
        # 备选方案：使用pypdf提取文本
        if HAS_PYPDF:
            return extract_pdf_text(file_bytes)
        return jsonify({'error': 'PDF转换支持未安装，请安装: pip install pdf2image'}), 500

    try:
        # 保存临时文件
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        try:
            # 转换PDF为图片
            images = convert_from_path(tmp_path, dpi=200)

            results = []
            for i, image in enumerate(images):
                if target_format.lower() in ['jpg', 'jpeg']:
                    if image.mode in ('RGBA', 'P'):
                        image = image.convert('RGB')

                output = io.BytesIO()
                image.save(output, format=target_format.upper())
                output.seek(0)

                results.append({
                    'page': i + 1,
                    'image': base64.b64encode(output.read()).decode('utf-8')
                })

            return jsonify({
                'success': True,
                'pages': results,
                'format': target_format
            })

        finally:
            os.unlink(tmp_path)

    except Exception as e:
        return jsonify({'error': f'PDF转换失败: {str(e)}'}), 500


def extract_pdf_text(file_bytes):
    """提取PDF文本内容"""
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        text_content = []

        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            text_content.append({
                'page': i + 1,
                'text': text
            })

        return jsonify({
            'success': True,
            'type': 'text',
            'pages': text_content
        })

    except Exception as e:
        return jsonify({'error': f'PDF文本提取失败: {str(e)}'}), 500


def convert_ofd(file_bytes, target_format='jpg'):
    """转换OFD格式（中国电子发票标准格式）"""
    # OFD格式转换需要专门的库
    # 这里提供一个框架，实际需要安装ofdrw等库

    try:
        # 尝试使用ofdrw
        try:
            from ofdrw.reader import OFDReader
            # OFD处理逻辑
            # 由于ofdrw较为复杂，这里返回提示
            return jsonify({
                'error': 'OFD格式需要额外配置。建议将OFD转为PDF后处理。',
                'suggestion': '可使用OFD阅读器转换为PDF'
            })
        except ImportError:
            return jsonify({
                'error': 'OFD格式支持未安装。',
                'suggestion': '可安装ofdrw库或手动转换OFD为PDF'
            }), 500

    except Exception as e:
        return jsonify({'error': f'OFD转换失败: {str(e)}'}), 500


def convert_docx(file_bytes, target_format='jpg'):
    """从Word文档中提取图片"""
    if not HAS_DOCX:
        return jsonify({'error': 'Word文档支持未安装，请安装: pip install python-docx'}), 500

    try:
        doc = Document(io.BytesIO(file_bytes))

        images = []
        image_count = 0

        # 提取文档中的图片
        for rel in doc.part.rels.values():
            if "image" in rel.target_ref:
                image_count += 1
                image_data = rel.target_part.blob

                # 转换为base64
                image_base64 = base64.b64encode(image_data).decode('utf-8')

                # 检测图片格式
                img = Image.open(io.BytesIO(image_data))
                images.append({
                    'index': image_count,
                    'image': image_base64,
                    'format': img.format.lower() if img.format else 'png',
                    'size': img.size
                })

        # 提取文本内容
        text_content = []
        for para in doc.paragraphs:
            if para.text.strip():
                text_content.append(para.text)

        return jsonify({
            'success': True,
            'images': images,
            'text': '\n'.join(text_content),
            'has_images': len(images) > 0
        })

    except Exception as e:
        return jsonify({'error': f'Word文档处理失败: {str(e)}'}), 500


def convert_xlsx(file_bytes, target_format='jpg'):
    """从Excel文档中提取数据"""
    if not HAS_OPENPYXL:
        return jsonify({'error': 'Excel支持未安装，请安装: pip install openpyxl'}), 500

    try:
        wb = load_workbook(io.BytesIO(file_bytes))

        sheets_data = []
        images = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # 提取表格数据
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                if any(cell for cell in row_data):  # 跳过空行
                    rows_data.append(row_data)

            if rows_data:
                sheets_data.append({
                    'name': sheet_name,
                    'rows': rows_data
                })

            # 提取图片
            if hasattr(sheet, '_images'):
                for idx, img in enumerate(sheet._images):
                    try:
                        img_data = img._data()
                        images.append({
                            'sheet': sheet_name,
                            'index': idx + 1,
                            'image': base64.b64encode(img_data).decode('utf-8')
                        })
                    except:
                        pass

        return jsonify({
            'success': True,
            'sheets': sheets_data,
            'images': images,
            'has_images': len(images) > 0
        })

    except Exception as e:
        return jsonify({'error': f'Excel处理失败: {str(e)}'}), 500


@app.route('/extract_images', methods=['POST'])
def extract_images():
    """
    从文档中提取图片
    """
    try:
        data = request.get_json()
        if not data or 'file' not in data:
            return jsonify({'error': '缺少文件数据'}), 400

        file_data = data['file']
        if file_data.startswith('data:'):
            file_data = file_data.split(',', 1)[1]

        filename = data.get('filename', 'unknown')
        file_bytes = base64.b64decode(file_data)

        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

        if ext == 'docx':
            return convert_docx(file_bytes)
        elif ext == 'xlsx':
            return convert_xlsx(file_bytes)
        elif ext == 'pdf':
            return convert_pdf(file_bytes, 'png')
        else:
            return jsonify({'error': f'不支持从 {ext} 格式提取图片'}), 400

    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("文件转换服务启动...")
    print("支持的格式:")
    print(f"  HEIC/HEIF: {'是' if HAS_HEIF else '否'}")
    print(f"  PDF: {'是' if HAS_PDF2IMAGE else '仅文本提取'}")
    print(f"  Word: {'是' if HAS_DOCX else '否'}")
    print(f"  Excel: {'是' if HAS_OPENPYXL else '否'}")
    print()
    print("API端点:")
    print("  GET  /health         - 健康检查")
    print("  POST /convert        - 文件格式转换")
    print("  POST /extract_images - 从文档提取图片")
    print()
    print("服务器端口: 5001")

    app.run(host='127.0.0.1', port=5001, debug=False)
