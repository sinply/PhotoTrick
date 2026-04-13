#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Command-line file converter for PhotoTrick
Usage: python file_converter_cli.py --action <action> --input <input> --output <output>
"""

import sys
import os
import argparse

# Check dependencies
HAS_PYMUPDF = False
HAS_HEIF = False

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    pass

try:
    from pillow_heif import register_heif_opener
    from PIL import Image
    register_heif_opener()
    HAS_HEIF = True
except ImportError:
    pass


def pdf_to_images(input_path, output_dir):
    """Convert PDF to images using PyMuPDF"""
    if not HAS_PYMUPDF:
        print("Error: PyMuPDF not installed. Run: pip install PyMuPDF", file=sys.stderr)
        return False

    try:
        os.makedirs(output_dir, exist_ok=True)
        doc = fitz.open(input_path)

        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            # Render at 200 DPI (zoom = 200/72)
            zoom = 200 / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)

            output_path = os.path.join(output_dir, f"page_{page_num+1:03d}.jpg")
            pix.save(output_path)
            print(f"Saved: {output_path}")

        doc.close()
        return True
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return False


def heic_to_jpg(input_path, output_path):
    """Convert HEIC to JPG"""
    if not HAS_HEIF:
        print("Error: pillow-heif not installed. Run: pip install pillow-heif", file=sys.stderr)
        return False

    try:
        from PIL import Image
        image = Image.open(input_path)
        if image.mode in ('RGBA', 'P'):
            image = image.convert('RGB')
        image.save(output_path, 'JPEG', quality=95)
        print(f"Saved: {output_path}")
        return True
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return False


def extract_docx_images(input_path, output_dir):
    """Extract images from DOCX"""
    try:
        from docx import Document
    except ImportError:
        print("Error: python-docx not installed. Run: pip install python-docx", file=sys.stderr)
        return False

    try:
        os.makedirs(output_dir, exist_ok=True)
        doc = Document(input_path)

        count = 0
        for rel in doc.part.rels.values():
            if "image" in rel.target_ref:
                count += 1
                image_data = rel.target_part.blob
                ext = rel.target_ref.split('.')[-1].lower()
                output_path = os.path.join(output_dir, f"image_{count:03d}.{ext}")
                with open(output_path, 'wb') as f:
                    f.write(image_data)
                print(f"Saved: {output_path}")

        return count > 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return False


def extract_xlsx_images(input_path, output_dir):
    """Extract images from XLSX"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return False

    try:
        os.makedirs(output_dir, exist_ok=True)
        wb = load_workbook(input_path)

        count = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if hasattr(sheet, '_images'):
                for idx, img in enumerate(sheet._images):
                    try:
                        count += 1
                        img_data = img._data()
                        output_path = os.path.join(output_dir, f"image_{count:03d}.png")
                        with open(output_path, 'wb') as f:
                            f.write(img_data)
                        print(f"Saved: {output_path}")
                    except Exception as e:
                        print(f"Warning: Could not extract image: {e}", file=sys.stderr)

        return count > 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return False


def ofd_to_images(input_path, output_dir):
    """Convert OFD to images - placeholder"""
    print("Error: OFD format not yet supported. Please convert OFD to PDF first.", file=sys.stderr)
    return False


def main():
    parser = argparse.ArgumentParser(description='File converter for PhotoTrick')
    parser.add_argument('--action', required=True, help='Action: pdf_to_images, heic_to_jpg, extract_docx_images, extract_xlsx_images, ofd_to_images')
    parser.add_argument('--input', required=True, help='Input file path')
    parser.add_argument('--output', required=True, help='Output file/directory path')

    args = parser.parse_args()

    actions = {
        'pdf_to_images': pdf_to_images,
        'heic_to_jpg': heic_to_jpg,
        'extract_docx_images': extract_docx_images,
        'extract_xlsx_images': extract_xlsx_images,
        'ofd_to_images': ofd_to_images
    }

    if args.action not in actions:
        print(f"Unknown action: {args.action}", file=sys.stderr)
        print(f"Available actions: {', '.join(actions.keys())}", file=sys.stderr)
        sys.exit(1)

    success = actions[args.action](args.input, args.output)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
